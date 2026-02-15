"""
Agentic Honey-Pot API v2.1.0
Scam Intelligence Platform — Problem Statement 2

This is the main entry point for the honeypot system. It receives suspected
scam messages, analyzes them, generates responses, extracts intelligence,
and reports findings to configured government portals.
"""
import os
from pathlib import Path
from dotenv import load_dotenv

# Load .env from repo root (parent of app/) or current dir
_env_file = Path(__file__).resolve().parent.parent / ".env"
if _env_file.exists():
    load_dotenv(_env_file)
else:
    load_dotenv()

from fastapi import FastAPI, Depends, HTTPException, Request, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.exceptions import RequestValidationError
from fastapi.responses import JSONResponse
from typing import Optional
import logging
import time
from datetime import datetime, timezone

from models import (
    HoneypotRequest,
    HoneypotResponse,
    SimulationRequest,
    SimulationResponse,
)
from auth import verify_api_key
from detector import detector
from extractor import extractor
from agent import agent
from memory import memory
from callback import send_final_callback, should_send_callback
from llm import llm_service, is_greeting_message
from db import db_service
from simulator import simulator
from intelligence import (
    IntelligenceRegistryService,
    PatternCorrelationService,
    classify_fraud_type,
    get_fraud_color,
    generate_detection_reasoning,
)

# Initialize intelligence services
intel_registry = IntelligenceRegistryService(db_service)
pattern_engine = PatternCorrelationService(db_service)

# ─── Structured Logging ──────────────────────────────────────────────────────
LOG_LEVEL = os.getenv("LOG_LEVEL", "INFO").upper()
logging.basicConfig(
    level=getattr(logging, LOG_LEVEL, logging.INFO),
    format="%(asctime)s  %(levelname)-8s  %(name)s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
# Silence noisy third-party loggers in production
for _noisy in ("httpcore", "httpx", "uvicorn.access", "watchfiles"):
    logging.getLogger(_noisy).setLevel(logging.WARNING)
logger = logging.getLogger("trusthoneypot")

# Create the FastAPI app
app = FastAPI(
    title="Agentic Honey-Pot API",
    description="Scam Detection & Intelligence Extraction Platform",
    version="2.1.0",
    docs_url="/docs" if os.getenv("ENVIRONMENT", "development") != "production" else None,
    redoc_url=None,
)

# CORS — restrict in production, wide open in dev
_default_origins = "https://trusthoneypot.tech,https://www.trusthoneypot.tech,http://localhost:5173,http://localhost:3000"
_allowed_origins = [o.strip() for o in os.getenv("CORS_ORIGINS", _default_origins).split(",") if o.strip()]
app.add_middleware(
    CORSMiddleware,
    allow_origins=_allowed_origins,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS", "PATCH"],
    allow_headers=["*"],
    expose_headers=["X-Process-Time"],
)


# ─── Production Middleware ────────────────────────────────────────────────────

@app.middleware("http")
async def add_timing_header(request: Request, call_next):
    """Add X-Process-Time header and log request timing."""
    # Let CORS middleware handle OPTIONS preflight directly
    if request.method == "OPTIONS":
        return await call_next(request)
    start = time.perf_counter()
    response = await call_next(request)
    elapsed_ms = (time.perf_counter() - start) * 1000
    response.headers["X-Process-Time"] = f"{elapsed_ms:.0f}ms"
    if elapsed_ms > 2000:
        logger.warning("slow_request  method=%s  path=%s  duration_ms=%.0f", request.method, request.url.path, elapsed_ms)
    return response


# ─── Simple in-memory rate limiter ────────────────────────────────────────────

_rate_store: dict = {}  # ip -> (count, window_start)
RATE_LIMIT = int(os.getenv("RATE_LIMIT_PER_MINUTE", "60"))

@app.middleware("http")
async def rate_limit_middleware(request: Request, call_next):
    """Basic per-IP rate limiting for production safety."""
    # Let CORS middleware handle OPTIONS preflight directly
    if request.method == "OPTIONS":
        return await call_next(request)
    client_ip = request.client.host if request.client else "unknown"
    now = time.time()
    window = 60  # 1 minute window
    
    entry = _rate_store.get(client_ip, (0, now))
    count, window_start = entry
    
    if now - window_start > window:
        # New window
        _rate_store[client_ip] = (1, now)
    elif count >= RATE_LIMIT:
        return JSONResponse(
            status_code=429,
            content={"detail": "Rate limit exceeded. Try again in a minute."},
        )
    else:
        _rate_store[client_ip] = (count + 1, window_start)
    
    # Cleanup old entries periodically (every 100 requests)
    if len(_rate_store) > 500:
        cutoff = now - window * 2
        _rate_store.update({
            k: v for k, v in _rate_store.items() if v[1] > cutoff
        })
    
    return await call_next(request)


@app.on_event("startup")
async def startup_event():
    """Log essential startup information and run initial cleanup."""
    memory.cleanup_stale_sessions()
    memory.enforce_limit()
    _env = os.getenv("ENVIRONMENT", "development")
    logger.info("="*60)
    logger.info("TrustHoneypot API v%s  env=%s", app.version, _env)
    logger.info("cors_origins=%s", _allowed_origins)
    logger.info("rate_limit=%d/min  log_level=%s", RATE_LIMIT, LOG_LEVEL)
    logger.info("routes  health=GET /  honeypot=POST /honeypot  docs=%s", "/docs" if app.docs_url else "disabled")
    logger.info("="*60)


@app.on_event("shutdown")
async def shutdown_event():
    """Clean up resources on shutdown."""
    await llm_service.close()
    logger.info("Graceful shutdown complete")


@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request: Request, exc: RequestValidationError):
    """Log validation errors clearly."""
    logger.error("validation_error  path=%s  detail=%s", request.url.path, exc.errors())
    
    return JSONResponse(
        status_code=422,
        content={
            "detail": exc.errors(),
            "message": "Invalid request payload. See detail for exact fields.",
        },
    )


@app.get("/")
async def health_check():
    """Simple health check - lets the platform know we're alive."""
    return {
        "status": "online",
        "service": "Agentic Honey-Pot API",
        "version": "2.1.0",
        "languages": ["en", "hi"],
    }


@app.post("/honeypot", response_model=HoneypotResponse)
async def process_message(
    request: HoneypotRequest,
    api_key: str = Depends(verify_api_key)
):
    """Process incoming scam messages and return analysis results."""
    try:
        import json
        
        session_id = request.sessionId
        current_message = request.message.text
        response_mode = getattr(request, "response_mode", None) or "rule_based"
        
        # Log full request body in one line
        request_dict = request.model_dump()
        logger.info("request  session=%s  mode=%s  message_len=%d", session_id[:8], response_mode, len(current_message))
        
        # Process conversation history for context
        # First, update agent's context awareness from history
        agent.process_conversation_history(session_id, request.conversationHistory)
        
        # Only process NEW history messages through detector/extractor
        # (avoid re-scoring the same messages on every request, which inflates risk scores)
        _history_key = f"_detector_hist_{session_id}"
        _already_processed = memory.sessions.get(session_id, {}).get("_detector_history_count", 0)
        _new_history = request.conversationHistory[_already_processed:]
        for hist_msg in _new_history:
            if hist_msg.sender == "scammer":
                detector.calculate_risk_score(hist_msg.text, session_id)
                extractor.extract(hist_msg.text, session_id)
        memory.create_session(session_id)
        memory.sessions[session_id]["_detector_history_count"] = len(request.conversationHistory)
        
        memory.add_message(session_id, "scammer", current_message)
        
        # Analyze current message
        risk_score, is_scam = detector.calculate_risk_score(current_message, session_id)
        detection_details = detector.get_detection_details(session_id)
        
        if is_scam and not memory.is_scam_confirmed(session_id):
            memory.mark_scam_confirmed(session_id)
        
        # Generate internal agent response (not returned to client)
        scam_confirmed = memory.is_scam_confirmed(session_id)
        # Use actual conversation length from history, not just server memory count
        msg_count = len(request.conversationHistory) + 1
        
        # Always generate a rule-based reply first (authority)
        agent_reply = agent.get_reply(session_id, current_message, msg_count, scam_confirmed)
        reply_source = "rule_based"
        
        # ──── LLM Response Generation ────────────────────────────────────────────
        # If LLM mode requested, we have two paths:
        # 1. Greeting-stage messages → use dedicated GREETING_LLM_PROMPT
        # 2. Normal scam messages → rephrase rule-based reply for naturalness
        if response_mode == "llm":
            logger.info("llm_mode  session=%s  enabled=%s  model=%s", session_id[:8], llm_service.enabled, llm_service.model_name)
            if not llm_service.enabled:
                llm_status = llm_service.get_status()
                logger.warning("llm_disabled  session=%s  status=%s", session_id[:8], json.dumps(llm_status))
                reply_source = "rule_based_fallback"
            elif is_greeting_message(current_message):
                # PATH 1: Greeting-stage detection
                # Use specialized greeting prompt that instructs LLM to be warm and polite,
                # not defensive or suspicious. This prevents the old behavior where LLM
                # would respond with "I'm not sure what this is about..." to simple "hi" messages.
                logger.info("llm_greeting  session=%s", session_id[:8])
                try:
                    llm_reply, llm_source = await llm_service.generate_greeting_reply(
                        scammer_message=current_message
                    )
                    agent_reply = llm_reply
                    reply_source = llm_source
                except Exception as llm_err:
                    logger.warning("llm_greeting_fallback  session=%s  error=%s", session_id[:8], llm_err)
                    reply_source = "rule_based_fallback"
            else:
                # PATH 2: Normal scam engagement
                # Rephrase the rule-based reply via LLM for more natural language
                strategy = agent.get_current_strategy(session_id)
                try:
                    llm_reply, llm_source = await llm_service.rephrase_reply(
                        strategy=strategy,
                        rule_reply=agent_reply,
                        scammer_message=current_message
                    )
                    agent_reply = llm_reply
                    reply_source = llm_source
                except Exception as llm_err:
                    logger.warning("llm_rephrase_fallback  session=%s  error=%s", session_id[:8], llm_err)
                    reply_source = "rule_based_fallback"
        
        memory.set_agent_response(session_id, agent_reply)
        memory.add_message(session_id, "agent", agent_reply)
        
        # Extract intelligence
        intelligence = extractor.extract(current_message, session_id)
        
        # Enrich suspiciousKeywords with detected categories for better analysis
        detected_categories = list(detection_details.triggered_categories)
        if detection_details.scam_type and detection_details.scam_type != "unknown":
            detected_categories.append(detection_details.scam_type)
        existing_keywords = intelligence.get("suspiciousKeywords", [])
        intelligence["suspiciousKeywords"] = list(set(existing_keywords + detected_categories))
        
        # Calculate metrics using conversation history length + 1
        total_messages = len(request.conversationHistory) + 1
        duration_seconds = memory.get_duration(session_id)
        
        # Generate notes with enhanced detection details (internal use only)
        if scam_confirmed:
            agent_notes = agent.generate_agent_notes(
                session_id, total_messages, intelligence, detection_details
            )
        else:
            agent_notes = agent.generate_monitoring_notes(session_id, total_messages)
        
        # Send callback if conditions met
        callback_sent = False
        logger.debug("callback_check  session=%s  scam=%s  msgs=%d", session_id[:8], scam_confirmed, total_messages)
        callback_eligible = should_send_callback(scam_confirmed, total_messages, intelligence)
        
        if callback_eligible:
            already_sent = memory.is_callback_sent(session_id)
            logger.info("callback_eligible  session=%s  already_sent=%s", session_id[:8], already_sent)
            if not already_sent:
                success = send_final_callback(session_id, total_messages, intelligence, agent_notes)
                if success:
                    memory.mark_callback_sent(session_id)
                    callback_sent = True
                    # Save callback record to DB
                    db_service.save_callback_record(
                        session_id=session_id,
                        status="sent",
                        payload_summary={
                            "totalMessages": total_messages,
                            "intelligenceCounts": {
                                "upiIds": len(intelligence.get("upiIds", [])),
                                "phoneNumbers": len(intelligence.get("phoneNumbers", [])),
                                "bankAccounts": len(intelligence.get("bankAccounts", [])),
                                "phishingLinks": len(intelligence.get("phishingLinks", [])),
                            }
                        },
                        intelligence=intelligence,
                    )
        
        # Save session summary to DB (every message updates the summary)
        intel_counts = {
            "upiIds": len(intelligence.get("upiIds", [])),
            "phoneNumbers": len(intelligence.get("phoneNumbers", [])),
            "bankAccounts": len(intelligence.get("bankAccounts", [])),
            "phishingLinks": len(intelligence.get("phishingLinks", [])),
            "emails": len(intelligence.get("emails", [])),
            "suspiciousKeywords": len(intelligence.get("suspiciousKeywords", [])),
        }
        tactics_list = list(agent._get_context(session_id).get("detected_tactics", set()))
        fraud_label = classify_fraud_type(detection_details.scam_type or "unknown")
        db_service.save_session_summary(
            session_id=session_id,
            scam_type=detection_details.scam_type or "unknown",
            risk_level=detection_details.risk_level or "minimal",
            confidence=detection_details.confidence,
            message_count=total_messages,
            scam_detected=scam_confirmed,
            intelligence_counts=intel_counts,
            tactics=tactics_list,
            response_mode=reply_source,
            callback_sent=memory.is_callback_sent(session_id),
            intelligence=intelligence,
            fraud_type=fraud_label,
        )
        
        # Register intelligence in the registry (v2.1)
        intel_registry.register_session_intelligence(
            session_id=session_id,
            intelligence=intelligence,
            risk_level=detection_details.risk_level or "minimal",
            confidence=detection_details.confidence,
        )
        
        # Register pattern for correlation (v2.1)
        all_identifiers = (
            intelligence.get("upiIds", []) +
            intelligence.get("phoneNumbers", []) +
            intelligence.get("bankAccounts", []) +
            intelligence.get("phishingLinks", []) +
            intelligence.get("emails", [])
        )
        correlation = pattern_engine.register_pattern(
            session_id=session_id,
            scam_type=detection_details.scam_type or "unknown",
            tactics=tactics_list,
            identifiers=all_identifiers,
            risk_level=detection_details.risk_level or "minimal",
            confidence=detection_details.confidence,
        )
        
        # Build response (status, reply, plus enriched metadata for UI)
        stage_info = agent.get_engagement_stage(session_id, msg_count, scam_confirmed, callback_sent)
        
        # ──── Greeting Stage Override ─────────────────────────────────────────────
        # During Stage 0 (Rapport Initialization), when we've only received a greeting,
        # we don't want to show false positive scam indicators in the UI.
        # Override risk/confidence to show clean monitoring state:
        # - risk_level = "minimal" (green)
        # - confidence = 0.0 (no detection yet)
        # - scam_type = "unknown" (not classified yet)
        # This gives a clean UI state until the scammer reveals their actual intent.
        is_greeting_stage = agent.is_in_greeting_stage(session_id)
        resp_risk_level = "minimal" if is_greeting_stage else detection_details.risk_level
        resp_confidence = 0.0 if is_greeting_stage else detection_details.confidence
        resp_scam_type = "unknown" if is_greeting_stage else (detection_details.scam_type or "unknown")
        
        # v2.1 enrichment: fraud classification + detection reasoning
        fraud_type = classify_fraud_type(resp_scam_type)
        fraud_color = get_fraud_color(fraud_type)
        reasoning = generate_detection_reasoning(
            scam_type=resp_scam_type,
            risk_level=resp_risk_level or "minimal",
            confidence=resp_confidence,
            tactics=tactics_list,
            intelligence_counts=intel_counts,
            pattern_match_count=correlation.get("match_count", 0),
            similarity_score=correlation.get("similarity_score", 0.0),
            recurring=correlation.get("recurring", False),
        )
        
        response = HoneypotResponse(
            status="success",
            reply=agent_reply,
            reply_source=reply_source,
            scam_detected=scam_confirmed,
            risk_score=0 if is_greeting_stage else risk_score,
            risk_level=resp_risk_level,
            confidence=resp_confidence,
            scam_type=resp_scam_type,
            scam_stage=stage_info["stage"],
            stage_info=stage_info,
            intelligence_counts=intel_counts,
            callback_sent=memory.is_callback_sent(session_id),
            fraud_type=fraud_type,
            fraud_color=fraud_color,
            detection_reasons=reasoning.get("reasons", []),
            pattern_similarity=correlation.get("similarity_score", 0.0),
        )
        
        # Internal logging - detection result, intelligence, notes, callback (not exposed in response)
        internal_log = {
            "scamDetected": scam_confirmed,
            "replySource": reply_source,
            "engagementMetrics": {
                "engagementDurationSeconds": duration_seconds,
                "totalMessagesExchanged": total_messages
            },
            "extractedIntelligence": {
                "bankAccounts": intelligence.get("bankAccounts", []),
                "upiIds": intelligence.get("upiIds", []),
                "phishingLinks": intelligence.get("phishingLinks", []),
                "phoneNumbers": intelligence.get("phoneNumbers", []),
                "suspiciousKeywords": intelligence.get("suspiciousKeywords", [])
            },
            "agentNotes": agent_notes
        }
        logger.debug("agent_internal  session=%s  data=%s", session_id[:8], json.dumps(internal_log, ensure_ascii=False))
        logger.info(
            "response  session=%s  risk=%s  confidence=%.2f  scam=%s  source=%s  callback=%s",
            session_id[:8], resp_risk_level, resp_confidence, scam_confirmed, reply_source,
            "sent" if callback_sent else ("eligible" if callback_eligible else "none"),
        )
        
        return response
        
    except Exception as e:
        logger.error("request_failed  error=%s", e, exc_info=True)
        raise HTTPException(status_code=500, detail=f"Internal error: {str(e)}")


def _get_scam_stage(msg_count: int, scam_confirmed: bool, callback_sent: bool) -> str:
    """Determine the scam lifecycle stage (legacy helper, prefer agent.get_engagement_stage)."""
    if callback_sent:
        return "intelligence_reported"
    if scam_confirmed and msg_count >= 5:
        return "deep_engagement"
    if scam_confirmed:
        return "scam_confirmed"
    if msg_count >= 2:
        return "monitoring"
    return "initial_contact"


# ─── Simulation Endpoints ─────────────────────────────────────────────────────

@app.get("/scenarios")
async def get_scenarios(api_key: str = Depends(verify_api_key)):
    """List available demo scam scenarios for auto-simulation."""
    return {"scenarios": simulator.get_scenarios()}


@app.get("/scenarios/{scenario_id}")
async def get_scenario_detail(scenario_id: str, api_key: str = Depends(verify_api_key)):
    """
    Get full scenario details including messages for step-by-step simulation.
    
    The frontend uses this to drive simulations message-by-message through
    the /honeypot endpoint, allowing real-time analysis updates after each step.
    """
    scenario = simulator.get_scenario(scenario_id)
    if not scenario:
        raise HTTPException(
            status_code=404,
            detail=f"Scenario '{scenario_id}' not found."
        )
    return {
        "id": scenario["id"],
        "name": scenario["name"],
        "description": scenario["description"],
        "language": scenario["language"],
        "scam_type": scenario["scam_type"],
        "difficulty": scenario["difficulty"],
        "messages": scenario["messages"],
    }


@app.post("/simulate", response_model=SimulationResponse)
async def run_simulation(
    request: SimulationRequest,
    api_key: str = Depends(verify_api_key)
):
    """
    Run an autonomous scam simulation with a predefined scenario.
    
    The simulation processes each scammer message through the full honeypot
    pipeline (detection → agent response → extraction → callback check).
    Returns the complete conversation with stage progression and analysis.
    """
    import json as _json
    
    scenario = simulator.get_scenario(request.scenario_id)
    if not scenario:
        raise HTTPException(
            status_code=404,
            detail=f"Scenario '{request.scenario_id}' not found. Use GET /scenarios for available options."
        )
    
    async def honeypot_handler(session_id, message_text, history, response_mode):
        """Internal handler that mirrors /honeypot logic for simulation."""
        from models import Message
        
        # Build message history objects
        history_msgs = [Message(sender=h["sender"], text=h["text"]) for h in history]
        
        # Process conversation history (only new messages)
        agent.process_conversation_history(session_id, history_msgs)
        
        _already = memory.sessions.get(session_id, {}).get("_detector_history_count", 0)
        for hist_msg in history_msgs[_already:]:
            if hist_msg.sender == "scammer":
                detector.calculate_risk_score(hist_msg.text, session_id)
                extractor.extract(hist_msg.text, session_id)
        memory.create_session(session_id)
        memory.sessions[session_id]["_detector_history_count"] = len(history_msgs)
        
        memory.add_message(session_id, "scammer", message_text)
        
        # Analyze current message
        risk_score, is_scam = detector.calculate_risk_score(message_text, session_id)
        detection_details = detector.get_detection_details(session_id)
        
        if is_scam and not memory.is_scam_confirmed(session_id):
            memory.mark_scam_confirmed(session_id)
        
        scam_confirmed = memory.is_scam_confirmed(session_id)
        msg_count = len(history) + 1
        
        # Generate agent reply
        agent_reply = agent.get_reply(session_id, message_text, msg_count, scam_confirmed)
        reply_source = "rule_based"
        
        # LLM generation/rephrasing if requested
        if response_mode == "llm":
            if is_greeting_message(message_text):
                # Greeting-stage: use dedicated greeting LLM prompt
                try:
                    llm_reply, llm_source = await llm_service.generate_greeting_reply(
                        scammer_message=message_text
                    )
                    agent_reply = llm_reply
                    reply_source = llm_source
                except Exception:
                    reply_source = "rule_based_fallback"
            else:
                # Normal scam engagement: rephrase rule-based reply via LLM
                strategy = agent.get_current_strategy(session_id)
                try:
                    llm_reply, llm_source = await llm_service.rephrase_reply(
                        strategy=strategy,
                        rule_reply=agent_reply,
                        scammer_message=message_text
                    )
                    agent_reply = llm_reply
                    reply_source = llm_source
                except Exception:
                    reply_source = "rule_based_fallback"
        
        memory.set_agent_response(session_id, agent_reply)
        memory.add_message(session_id, "agent", agent_reply)
        
        # Extract intelligence
        intelligence = extractor.extract(message_text, session_id)
        intel_counts = {
            "upiIds": len(intelligence.get("upiIds", [])),
            "phoneNumbers": len(intelligence.get("phoneNumbers", [])),
            "bankAccounts": len(intelligence.get("bankAccounts", [])),
            "phishingLinks": len(intelligence.get("phishingLinks", [])),
            "emails": len(intelligence.get("emails", [])),
            "suspiciousKeywords": len(intelligence.get("suspiciousKeywords", [])),
        }
        
        # Check callback
        total_messages = len(history) + 1
        callback_sent = False
        callback_eligible = should_send_callback(scam_confirmed, total_messages, intelligence)
        if callback_eligible and not memory.is_callback_sent(session_id):
            agent_notes = agent.generate_agent_notes(session_id, total_messages, intelligence, detection_details)
            success = send_final_callback(session_id, total_messages, intelligence, agent_notes)
            if success:
                memory.mark_callback_sent(session_id)
                callback_sent = True
                # Save callback record to DB (was missing in simulation handler!)
                db_service.save_callback_record(
                    session_id=session_id,
                    status="sent",
                    payload_summary={
                        "totalMessages": total_messages,
                        "intelligenceCounts": intel_counts,
                        "source": "simulation",
                    },
                    intelligence=intelligence,
                )
                logger.info("sim_callback_saved  session=%s", session_id[:8])
        
        # Save session summary to DB (was missing in simulation handler!)
        tactics_list = list(agent._get_context(session_id).get("detected_tactics", set()))
        fraud_label = classify_fraud_type(detection_details.scam_type or "unknown")
        db_service.save_session_summary(
            session_id=session_id,
            scam_type=detection_details.scam_type or "unknown",
            risk_level=detection_details.risk_level or "minimal",
            confidence=detection_details.confidence,
            message_count=total_messages,
            scam_detected=scam_confirmed,
            intelligence_counts=intel_counts,
            tactics=tactics_list,
            response_mode=reply_source,
            callback_sent=memory.is_callback_sent(session_id),
            intelligence=intelligence,
            fraud_type=fraud_label,
        )
        
        stage_info = agent.get_engagement_stage(session_id, msg_count, scam_confirmed, callback_sent or memory.is_callback_sent(session_id))
        
        return {
            "reply": agent_reply,
            "reply_source": reply_source,
            "scam_detected": scam_confirmed,
            "risk_score": risk_score,
            "risk_level": detection_details.risk_level,
            "confidence": detection_details.confidence,
            "scam_type": detection_details.scam_type or "unknown",
            "scam_stage": stage_info["stage"],
            "stage_info": stage_info,
            "intelligence_counts": intel_counts,
            "callback_sent": callback_sent or memory.is_callback_sent(session_id),
        }
    
    logger.info("simulation_start  scenario=%s  mode=%s", request.scenario_id, request.response_mode)
    
    result = await simulator.run_simulation(
        scenario_id=request.scenario_id,
        response_mode=request.response_mode,
        honeypot_handler=honeypot_handler,
    )
    
    if "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])
    
    return result


# ─── Read-Only Endpoints for UI ──────────────────────────────────────────────

@app.get("/sessions")
async def get_sessions(
    limit: int = Query(default=50, le=200),
    api_key: str = Depends(verify_api_key)
):
    """Get session summaries (no raw chats). For UI dashboard."""
    summaries = db_service.get_session_summaries(limit=limit)
    # Normalize DB records (camelCase) to snake_case for frontend
    normalized = []
    for s in summaries:
        scam_type = s.get("scamType", "unknown")
        normalized.append({
            "session_id": s.get("sessionId", ""),
            "scam_type": scam_type,
            "fraud_type": classify_fraud_type(scam_type),
            "fraud_color": get_fraud_color(classify_fraud_type(scam_type)),
            "risk_level": s.get("riskLevel", "minimal"),
            "confidence": s.get("confidence", 0.0),
            "message_count": s.get("messageCount", 0),
            "scam_confirmed": s.get("scamDetected", False),
            "intelligence_counts": s.get("intelligenceTypes", {}),
            "callback_sent": s.get("callbackSent", False),
            "response_mode": s.get("responseMode", "rule_based"),
            "tactics": s.get("tactics", []),
        })
    # Also include in-memory sessions not yet in DB
    for sid, session in memory.sessions.items():
        det = detector.get_detection_details(sid)
        normalized.append({
            "session_id": sid,
            "scam_confirmed": session.get("scamConfirmed", False),
            "message_count": session.get("messageCount", 0),
            "callback_sent": session.get("callbackSent", False),
            "risk_level": getattr(det, "risk_level", "minimal"),
            "scam_type": getattr(det, "scam_type", "unknown"),
            "confidence": getattr(det, "confidence", 0.0),
            "intelligence_counts": {},
            "active": True,
        })
    return {"sessions": normalized}


@app.get("/sessions/{session_id}")
async def get_session_detail(
    session_id: str,
    api_key: str = Depends(verify_api_key)
):
    """Get a single session's summary."""
    summary = db_service.get_session_summary(session_id)
    det = detector.get_detection_details(session_id)
    intel = extractor.get_intelligence_summary(session_id)
    return {
        "summary": summary,
        "detection": {
            "riskLevel": getattr(det, "risk_level", "minimal"),
            "confidence": getattr(det, "confidence", 0.0),
            "scamType": getattr(det, "scam_type", "unknown"),
            "riskScore": getattr(det, "total_score", 0),
        },
        "intelligenceCounts": intel,
    }


@app.get("/patterns")
async def get_patterns(api_key: str = Depends(verify_api_key)):
    """Get aggregated scam patterns and learning data."""
    return db_service.get_patterns()


@app.get("/callbacks")
async def get_callbacks(
    limit: int = Query(default=50, le=200),
    api_key: str = Depends(verify_api_key)
):
    """Get callback records for UI."""
    records = db_service.get_callback_records(limit=limit)
    # Normalize camelCase -> snake_case for frontend
    normalized = []
    for r in records:
        # Convert timestamp to ISO format string with timezone
        ts = r.get("timestamp")
        if ts:
            if hasattr(ts, 'isoformat'):
                ts = ts.isoformat()
            # Ensure it ends with Z for UTC
            if isinstance(ts, str) and not ts.endswith('Z') and '+' not in ts:
                ts = ts + 'Z'
        # Lookup session scam type for fraud label
        session_summary = db_service.get_session_summary(r.get("sessionId", ""))
        scam_type = session_summary.get("scamType", "unknown") if session_summary else "unknown"
        normalized.append({
            "session_id": r.get("sessionId", ""),
            "status": r.get("status", "unknown"),
            "payload_summary": r.get("payloadSummary", None),
            "intelligence": r.get("intelligence", None),
            "fraud_type": classify_fraud_type(scam_type),
            "fraud_color": get_fraud_color(classify_fraud_type(scam_type)),
            "timestamp": ts,
        })
    return {"callbacks": normalized}


# ─── Intelligence Registry Endpoints (v2.1) ──────────────────────────────────

@app.get("/intelligence/registry")
async def get_intelligence_registry(
    type: Optional[str] = Query(default=None, description="Filter by type: upi, phone, bank_account, link, email"),
    risk: Optional[str] = Query(default=None, description="Filter by risk level"),
    limit: int = Query(default=100, le=500),
    api_key: str = Depends(verify_api_key),
):
    """Get intelligence registry entries with optional filters."""
    db_type = _TYPE_FILTER_MAP.get(type) if type else None
    entries = intel_registry.get_registry(id_type=db_type, risk_level=risk, limit=limit)
    stats = intel_registry.get_registry_stats()
    identifiers = [_normalize_registry_entry(e) for e in entries]
    return {
        "identifiers": identifiers,
        "stats": {
            "total_identifiers": stats.get("total", 0),
            "by_type": stats.get("by_type", {}),
            "by_risk": stats.get("by_risk", {}),
            "recurring_count": sum(1 for i in identifiers if i["is_recurring"]),
        },
    }


@app.get("/intelligence/registry/{identifier}")
async def get_identifier_detail(
    identifier: str,
    api_key: str = Depends(verify_api_key),
):
    """Get detailed info for a specific identifier."""
    detail = intel_registry.get_identifier_detail(identifier)
    if not detail:
        raise HTTPException(status_code=404, detail="Identifier not found")
    occ = detail.get("occurrences", 1)
    sessions = detail.get("sessions", [])
    # Lookup fraud types from associated sessions
    fraud_types = set()
    for sid in sessions[:20]:
        s = db_service.get_session_summary(sid)
        if s:
            fraud_types.add(classify_fraud_type(s.get("scamType", "unknown")))
    return {
        "identifier": detail.get("value", identifier),
        "masked_value": detail.get("masked", identifier),
        "type": _TYPE_DISPLAY_MAP.get(detail.get("type", ""), detail.get("type", "").lower()),
        "risk_level": detail.get("riskLevel", ""),
        "confidence": detail.get("confidence", 0),
        "frequency": occ,
        "is_recurring": occ > 1,
        "first_seen": detail.get("firstSeen", ""),
        "last_seen": detail.get("lastSeen", ""),
        "associated_sessions": sessions,
        "fraud_types": sorted(fraud_types),
    }


# Type mappings for intelligence registry normalization
_TYPE_FILTER_MAP = {"upi": "UPI", "phone": "Phone", "bank_account": "Bank", "link": "Link", "email": "Email"}
_TYPE_DISPLAY_MAP = {"UPI": "upi", "Phone": "phone", "Bank": "bank_account", "Link": "link", "Email": "email"}


def _normalize_registry_entry(e: dict) -> dict:
    """Normalize a registry entry from DB camelCase to frontend snake_case."""
    occ = e.get("occurrences", 1)
    return {
        "identifier": e.get("value", ""),
        "masked_value": e.get("masked", e.get("value", "")),
        "type": _TYPE_DISPLAY_MAP.get(e.get("type", ""), e.get("type", "").lower()),
        "risk_level": e.get("riskLevel", ""),
        "confidence": e.get("confidence", 0),
        "frequency": occ,
        "is_recurring": occ > 1,
        "first_seen": e.get("firstSeen", ""),
        "last_seen": e.get("lastSeen", ""),
        "sessions": e.get("sessions", []),
    }


@app.get("/intelligence/patterns")
async def get_pattern_correlation(api_key: str = Depends(verify_api_key)):
    """Get pattern correlation statistics."""
    raw = pattern_engine.get_pattern_stats()
    patterns = []
    for p in raw.get("top_patterns", []):
        cnt = p.get("count", 0)
        patterns.append({
            "scam_type": p.get("scam_type", "unknown"),
            "occurrence_count": cnt,
            "similarity_score": 1.0 if cnt > 1 else 0.0,
            "hash": p.get("hash", ""),
            "tactics": p.get("tactics", []),
        })
    return {
        "patterns": patterns,
        "stats": {
            "total_patterns": raw.get("total_patterns", 0),
            "recurring_patterns": raw.get("recurring_count", 0),
            "avg_similarity": raw.get("avg_similarity", 0.0),
            "unique_scam_types": raw.get("unique_scam_types", 0),
        },
    }


@app.get("/sessions/{session_id}/analysis")
async def get_session_analysis(
    session_id: str,
    api_key: str = Depends(verify_api_key),
):
    """Get enhanced session analysis with detection reasoning (v2.1)."""
    summary = db_service.get_session_summary(session_id)
    det = detector.get_detection_details(session_id)
    intel = extractor.get_intelligence_summary(session_id)
    
    # Pattern correlation
    tactics = summary.get("tactics", []) if summary else []
    scam_type = summary.get("scamType", "unknown") if summary else getattr(det, "scam_type", "unknown")
    
    # Get correlation info from pattern registry
    correlation_info = {"match_count": 0, "recurring": False, "similarity_score": 0.0}
    if db_service.enabled and db_service.db is not None:
        try:
            pattern_doc = db_service.db.pattern_registry.find_one(
                {"sessionId": session_id}, {"_id": 0}
            )
            if pattern_doc:
                from pymongo import DESCENDING
                similar = db_service.db.pattern_registry.count_documents(
                    {"patternHash": pattern_doc.get("patternHash"), "sessionId": {"$ne": session_id}}
                )
                correlation_info["match_count"] = similar
                correlation_info["recurring"] = similar > 0
                correlation_info["similarity_score"] = 1.0 if similar > 0 else 0.0
                correlation_info["pattern_hash"] = pattern_doc.get("patternHash")
        except Exception:
            pass
    
    reasoning = generate_detection_reasoning(
        scam_type=scam_type,
        risk_level=getattr(det, "risk_level", "minimal"),
        confidence=getattr(det, "confidence", 0.0),
        tactics=tactics,
        intelligence_counts=intel,
        pattern_match_count=correlation_info["match_count"],
        similarity_score=correlation_info["similarity_score"],
        recurring=correlation_info["recurring"],
    )
    
    return {
        "session_id": session_id,
        "summary": summary,
        "detection": {
            "riskLevel": getattr(det, "risk_level", "minimal"),
            "confidence": getattr(det, "confidence", 0.0),
            "scamType": scam_type,
            "riskScore": getattr(det, "total_score", 0),
        },
        "intelligenceCounts": intel,
        "reasoning": reasoning,
        "correlation": correlation_info,
    }


@app.get("/intelligence/export")
async def export_intelligence_excel(
    type: Optional[str] = Query(default=None),
    risk: Optional[str] = Query(default=None),
    date_from: Optional[str] = Query(default=None, description="ISO date YYYY-MM-DD"),
    date_to: Optional[str] = Query(default=None, description="ISO date YYYY-MM-DD"),
    api_key: str = Depends(verify_api_key),
):
    """Export intelligence registry to Excel (.xlsx)."""
    from fastapi.responses import StreamingResponse
    import io
    
    try:
        from openpyxl import Workbook  # type: ignore[import-untyped]
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # type: ignore[import-untyped]
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    db_type = _TYPE_FILTER_MAP.get(type) if type else None
    entries = intel_registry.get_registry(id_type=db_type, risk_level=risk, limit=5000)
    
    # Apply date filtering
    if date_from or date_to:
        filtered = []
        for e in entries:
            first_seen = e.get("firstSeen", "")
            if isinstance(first_seen, str):
                date_part = first_seen[:10]
            else:
                date_part = ""
            if date_from and date_part < date_from:
                continue
            if date_to and date_part > date_to:
                continue
            filtered.append(e)
        entries = filtered
    
    wb = Workbook()
    ws = wb.active
    assert ws is not None, "Workbook must have an active sheet"
    ws.title = "Intelligence Registry"
    
    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    
    headers = ["Identifier", "Type", "Risk Level", "Confidence %", "Occurrences", "First Seen", "Last Seen", "Sessions"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Risk color fills
    risk_fills = {
        "critical": PatternFill("solid", fgColor="FFCCCC"),
        "high": PatternFill("solid", fgColor="FFE0CC"),
        "medium": PatternFill("solid", fgColor="FFFACC"),
        "low": PatternFill("solid", fgColor="CCFFCC"),
    }
    
    for row_idx, entry in enumerate(entries, 2):
        ws.cell(row=row_idx, column=1, value=entry.get("masked", entry.get("value", ""))).border = thin_border
        ws.cell(row=row_idx, column=2, value=entry.get("type", "")).border = thin_border
        risk_cell = ws.cell(row=row_idx, column=3, value=(entry.get("riskLevel", "")).upper())
        risk_cell.border = thin_border
        risk_cell.fill = risk_fills.get(entry.get("riskLevel", ""), PatternFill())
        conf_cell = ws.cell(row=row_idx, column=4, value=round(entry.get("confidence", 0) * 100, 1))
        conf_cell.border = thin_border
        conf_cell.alignment = Alignment(horizontal="center")
        occ_cell = ws.cell(row=row_idx, column=5, value=entry.get("occurrences", 0))
        occ_cell.border = thin_border
        occ_cell.alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=6, value=str(entry.get("firstSeen", ""))[:19]).border = thin_border
        ws.cell(row=row_idx, column=7, value=str(entry.get("lastSeen", ""))[:19]).border = thin_border
        sessions = entry.get("sessions", [])
        ws.cell(row=row_idx, column=8, value=", ".join(s[:8] for s in sessions)).border = thin_border
    
    # Auto-width
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 40)  # type: ignore[union-attr]
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"trusthoneypot_intelligence_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.post("/intelligence/backfill")
async def backfill_intelligence(api_key: str = Depends(verify_api_key)):
    """Backfill intelligence registry + pattern registry from existing session summaries.
    
    Useful after fixing the DB boolean check bug that prevented registration.
    """
    if not db_service.enabled:
        raise HTTPException(status_code=503, detail="Database not available")
    
    summaries = db_service.get_session_summaries(limit=500)
    backfilled = 0
    for s in summaries:
        intel = s.get("intelligence", {})
        if not intel:
            continue
        session_id = s.get("sessionId", "")
        risk_level = s.get("riskLevel", "minimal")
        confidence = s.get("confidence", 0.0)
        scam_type = s.get("scamType", "unknown")
        tactics = s.get("tactics", [])
        
        # Register identifiers
        intel_registry.register_session_intelligence(
            session_id=session_id,
            intelligence=intel,
            risk_level=risk_level,
            confidence=confidence,
        )
        # Register pattern
        all_ids = (
            intel.get("upiIds", []) +
            intel.get("phoneNumbers", []) +
            intel.get("bankAccounts", []) +
            intel.get("phishingLinks", []) +
            intel.get("emails", [])
        )
        pattern_engine.register_pattern(
            session_id=session_id,
            scam_type=scam_type,
            tactics=tactics,
            identifiers=all_ids,
            risk_level=risk_level,
            confidence=confidence,
        )
        backfilled += 1
    
    logger.info("intelligence_backfill  sessions=%d", backfilled)
    return {"backfilled": backfilled, "message": f"Processed {backfilled} sessions"}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000, log_level="info")
